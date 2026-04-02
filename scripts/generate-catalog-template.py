#!/usr/bin/env python3
"""
Generate the GullyBite catalog upload XLSX template.
Matches exactly what backend/src/routes/restaurant.js upload handlers expect.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'public', 'templates', 'catalog-upload-template.xlsx')

# ── Colors ──
GREEN_HDR = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
VEG_GREEN = Font(color='2E7D32', bold=True)
NONVEG_RED = Font(color='C62828', bold=True)
EGG_YELLOW = Font(color='F9A825', bold=True)
JAIN_BLUE = Font(color='1565C0', bold=True)
VEGAN_PURPLE = Font(color='6A1B9A', bold=True)
THIN_BORDER = Border(
    bottom=Side(style='thin', color='E0E0E0'),
)

# ── Column definitions (matches META_COLUMN_ALIASES keys) ──
# (header, width, format_hint, required)
COLUMNS = [
    ('title',                        25, 'e.g., Chicken Biryani',                          True),
    ('description',                  45, 'Short description (10+ chars recommended)',       False),
    ('price',                        14, '199 or 199.00 or 199.00 INR',                    True),
    ('sale_price',                   14, 'Discounted price (same format as price)',         False),
    ('availability',                 16, 'in stock / out of stock',                         False),
    ('size',                         14, 'e.g., Half / Full / Regular / Large',             False),
    ('item_group_id',                22, 'Auto-generated if size is set. Same for variants', False),
    ('category',                     18, 'e.g., Starters, Main Course, Breads',             False),
    ('food_type',                    12, 'veg / non_veg / egg / vegan',                     False),
    ('is_bestseller',                12, 'true / false',                                    False),
    ('image_link',                   35, 'https://your-image-url.jpg',                      False),
    ('brand',                        18, 'Restaurant name (auto-filled if empty)',           False),
    ('link',                         30, 'Product page URL (auto-generated)',                False),
    ('condition',                    12, 'Always: new',                                     False),
    ('google_product_category',      40, 'Food, Beverages & Tobacco > Food Items',          False),
    ('fb_product_category',          35, 'Food & Beverages > Prepared Food',                False),
    ('product_tags[0]',              16, 'Veg / Non-Veg / Egg / Jain / Vegan',             False),
    ('product_tags[1]',              16, 'Sub-category tag',                                False),
    ('branch',                       16, 'Branch slug for multi-branch upload',             False),
    ('custom_label_0',               16, 'Custom label 0',                                  False),
    ('custom_label_1',               16, 'Custom label 1',                                  False),
    ('custom_label_2',               16, 'Area / locality',                                 False),
    ('custom_label_3',               16, 'Branch slug (alt to "branch" column)',            False),
    ('custom_label_4',               16, 'Custom label 4',                                  False),
    ('id',                           25, 'retailer_id (auto-generated, leave empty)',       False),
    ('sale_price_effective_date',     22, 'ISO date range',                                  False),
    ('quantity_to_sell_on_facebook',  12, 'Stock qty (optional)',                            False),
    ('gtin',                         16, 'Barcode (optional)',                               False),
]

# ── Dummy menu items ──
ITEMS = [
    # Veg Starters
    {'title': 'Paneer Tikka', 'description': 'Marinated cottage cheese grilled in tandoor with bell peppers', 'price': '249', 'category': 'Starters', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock', 'size': '', 'item_group_id': ''},
    {'title': 'Veg Spring Roll', 'description': 'Crispy rolls stuffed with mixed vegetables', 'price': '179', 'category': 'Starters', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Hara Bhara Kebab', 'description': 'Spinach and green pea patties with mint chutney', 'price': '149', 'size': 'Half', 'item_group_id': 'koramangala-hara-bhara-kebab', 'category': 'Starters', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Hara Bhara Kebab', 'description': 'Spinach and green pea patties with mint chutney', 'price': '249', 'size': 'Full', 'item_group_id': 'koramangala-hara-bhara-kebab', 'category': 'Starters', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    # Non-Veg Starters
    {'title': 'Chicken 65', 'description': 'Spicy deep-fried chicken with curry leaves and chillies', 'price': '289', 'category': 'Starters', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Mutton Seekh Kebab', 'description': 'Minced lamb skewers grilled over charcoal', 'price': '349', 'category': 'Starters', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Fish Amritsari', 'description': 'Batter-fried fish fillets with tangy dip', 'price': '329', 'category': 'Starters', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    # Main Course
    {'title': 'Dal Makhani', 'description': 'Slow-cooked black lentils in creamy tomato gravy', 'price': '219', 'category': 'Main Course', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Paneer Butter Masala', 'description': 'Cottage cheese cubes in rich buttery tomato sauce', 'price': '269', 'category': 'Main Course', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Butter Chicken', 'description': 'Tender chicken in creamy tomato-butter gravy', 'price': '199', 'size': 'Half', 'item_group_id': 'indiranagar-butter-chicken', 'category': 'Main Course', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Butter Chicken', 'description': 'Tender chicken in creamy tomato-butter gravy', 'price': '349', 'size': 'Full', 'item_group_id': 'indiranagar-butter-chicken', 'category': 'Main Course', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Mutton Rogan Josh', 'description': 'Kashmiri-style lamb curry with aromatic spices', 'price': '399', 'category': 'Main Course', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'whitefield', 'availability': 'in stock'},
    # Biryani
    {'title': 'Chicken Biryani', 'description': 'Hyderabadi-style dum biryani with tender chicken', 'price': '249', 'size': 'Single', 'item_group_id': 'whitefield-chicken-biryani', 'category': 'Biryani', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'whitefield', 'availability': 'in stock'},
    {'title': 'Chicken Biryani', 'description': 'Hyderabadi-style dum biryani with tender chicken', 'price': '599', 'size': 'Family', 'item_group_id': 'whitefield-chicken-biryani', 'category': 'Biryani', 'food_type': 'non_veg', 'product_tags[0]': 'Non-Veg', 'branch': 'whitefield', 'availability': 'in stock'},
    {'title': 'Veg Biryani', 'description': 'Fragrant basmati rice with mixed vegetables and saffron', 'price': '199', 'category': 'Biryani', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'whitefield', 'availability': 'in stock'},
    # Breads
    {'title': 'Butter Naan', 'description': 'Soft leavened bread brushed with butter', 'price': '49', 'category': 'Breads', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Garlic Naan', 'description': 'Naan topped with fresh garlic and coriander', 'price': '59', 'category': 'Breads', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'koramangala', 'availability': 'in stock'},
    # Beverages
    {'title': 'Mango Lassi', 'description': 'Chilled yogurt smoothie with Alphonso mango', 'price': '99', 'size': 'Regular', 'item_group_id': 'indiranagar-mango-lassi', 'category': 'Beverages', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Mango Lassi', 'description': 'Chilled yogurt smoothie with Alphonso mango', 'price': '149', 'size': 'Large', 'item_group_id': 'indiranagar-mango-lassi', 'category': 'Beverages', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'indiranagar', 'availability': 'in stock'},
    {'title': 'Masala Chai', 'description': 'Traditional Indian spiced tea with ginger and cardamom', 'price': '49', 'category': 'Beverages', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'branch': 'whitefield', 'availability': 'in stock'},
    # Special diet
    {'title': 'Jain Paneer Tikka', 'description': 'No onion no garlic paneer tikka with mild spices', 'price': '279', 'category': 'Starters', 'food_type': 'veg', 'product_tags[0]': 'Jain', 'product_tags[1]': 'No Onion No Garlic', 'branch': 'koramangala', 'availability': 'in stock'},
    {'title': 'Vegan Buddha Bowl', 'description': 'Quinoa, roasted chickpeas, avocado, tahini dressing', 'price': '329', 'category': 'Main Course', 'food_type': 'vegan', 'product_tags[0]': 'Vegan', 'product_tags[1]': 'Healthy', 'branch': 'whitefield', 'availability': 'in stock'},
    # Egg
    {'title': 'Egg Biryani', 'description': 'Aromatic rice layered with spiced boiled eggs', 'price': '179', 'category': 'Biryani', 'food_type': 'egg', 'product_tags[0]': 'Egg', 'branch': 'whitefield', 'availability': 'in stock'},
    # Combo
    {'title': 'Thali Meal', 'description': 'Complete meal: dal, sabzi, rice, 2 roti, raita, papad, sweet', 'price': '299', 'category': 'Combos', 'food_type': 'veg', 'product_tags[0]': 'Veg', 'product_tags[1]': 'Combo', 'branch': 'koramangala', 'availability': 'in stock', 'is_bestseller': 'true'},
]

TAG_FONTS = {
    'Veg': VEG_GREEN, 'Non-Veg': NONVEG_RED, 'Egg': EGG_YELLOW,
    'Jain': JAIN_BLUE, 'Vegan': VEGAN_PURPLE,
}


def build_menu_sheet(wb):
    ws = wb.active
    ws.title = 'Menu Template'
    headers = [c[0] for c in COLUMNS]
    widths = [c[1] for c in COLUMNS]
    hints = [c[2] for c in COLUMNS]
    required = [c[3] for c in COLUMNS]

    # Header row
    for col_idx, (hdr, req) in enumerate(zip(headers, required), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr + (' *' if req else ''))
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = GREEN_HDR
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format hint row
    for col_idx, hint in enumerate(hints, 1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = Font(italic=True, color='757575', size=9)
        cell.fill = LIGHT_GRAY
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Data rows
    for row_idx, item in enumerate(ITEMS, 3):
        bg = LIGHT_GREEN if (row_idx % 2 == 0) else WHITE
        for col_idx, hdr in enumerate(headers, 1):
            val = item.get(hdr, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = bg
            cell.border = THIN_BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center')
            # Color-code diet tags
            if hdr == 'product_tags[0]' and val in TAG_FONTS:
                cell.font = TAG_FONTS[val]
            if hdr == 'food_type':
                fmap = {'veg': VEG_GREEN, 'non_veg': NONVEG_RED, 'egg': EGG_YELLOW, 'vegan': VEGAN_PURPLE}
                if val in fmap:
                    cell.font = fmap[val]

    # Column widths
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header + hint
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'


def build_instructions_sheet(wb):
    ws = wb.create_sheet('Instructions')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40

    hdr_font = Font(bold=True, size=13, color='1B5E20')
    sub_font = Font(bold=True, size=11)
    body_font = Font(size=10)

    r = 1
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='GullyBite Catalog Upload Template - Instructions').font = hdr_font
    r += 2

    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='Quick Start').font = sub_font
    r += 1
    for line in [
        '1. Fill in the "Menu Template" sheet. Required columns are marked with *.',
        '2. Upload this file on your GullyBite dashboard: Menu > + Add > Upload File.',
        '3. Map any unrecognized columns in the mapper dialog, then confirm.',
        '4. Items sync to your WhatsApp catalog automatically after upload.',
    ]:
        ws.cell(row=r, column=2, value=line).font = body_font
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='Column Reference').font = sub_font
    r += 1
    for hdr_label, col_letter in [('Column', 'B'), ('Required', 'C'), ('Format', 'D'), ('Example', 'E'), ('Description', 'F')]:
        ws.cell(row=r, column=ord(col_letter) - 64, value=hdr_label).font = Font(bold=True, size=10)
    r += 1
    ref_data = [
        ('title *', 'Yes', 'Text', 'Chicken Biryani', 'Item name shown to customers'),
        ('description', 'No', 'Text (10+ chars)', 'Hyderabadi dum biryani', 'Short description for catalog'),
        ('price *', 'Yes', 'Number or "N INR"', '249 or 249.00 INR', 'Price in INR (rupees, not paise)'),
        ('sale_price', 'No', 'Same as price', '199', 'Discounted price if on sale'),
        ('availability', 'No', 'in stock / out of stock', 'in stock', 'Default: in stock'),
        ('size', 'No', 'Text', 'Half / Full / Single', 'Creates variant. Same-name items auto-group'),
        ('item_group_id', 'No', 'branch-slug-item-slug', '', 'Auto-generated when size is set. Leave empty.'),
        ('category', 'No', 'Text', 'Starters', 'Menu category. Auto-created if new.'),
        ('food_type', 'No', 'veg / non_veg / egg / vegan', 'non_veg', 'Diet type. Default: veg'),
        ('is_bestseller', 'No', 'true / false', 'true', 'Highlight as bestseller'),
        ('image_link', 'No', 'HTTPS URL', 'https://...', 'Product image (JPEG/PNG, 600px+)'),
        ('brand', 'No', 'Text', 'Beyond Snacks', 'Auto-filled with restaurant name'),
        ('branch', 'No', 'Text/slug', 'koramangala', 'For multi-branch upload. Auto-creates new branches.'),
        ('product_tags[0]', 'No', 'Tag value', 'Veg / Non-Veg / Egg', 'Primary diet/category tag'),
        ('product_tags[1]', 'No', 'Tag value', 'Healthy / Combo', 'Secondary tag'),
        ('custom_label_3', 'No', 'branch slug', 'koramangala', 'Alternative to "branch" column'),
    ]
    for row_data in ref_data:
        for ci, val in enumerate(row_data, 2):
            ws.cell(row=r, column=ci, value=val).font = body_font
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='How Variants Work').font = sub_font
    r += 1
    for line in [
        'If an item comes in multiple sizes, create ONE ROW PER SIZE:',
        '',
        '  title              | price | size   | item_group_id',
        '  Chicken Biryani    | 249   | Single | (leave empty)',
        '  Chicken Biryani    | 599   | Family | (leave empty)',
        '',
        'The system auto-groups items with the same title as variants.',
        'item_group_id is auto-generated from branch slug + item name.',
        'You can set it manually to force grouping: e.g., "koramangala-chicken-biryani"',
    ]:
        ws.cell(row=r, column=2, value=line).font = body_font
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='How Branches Work').font = sub_font
    r += 1
    for line in [
        'For multi-branch restaurants, add a "branch" column (or use custom_label_3).',
        'Use lowercase slugs with hyphens: koramangala, hsr-layout, indiranagar.',
        'New branches are auto-created during upload (you must set GPS coordinates later).',
        'If no branch column is present, all items go to the selected branch.',
    ]:
        ws.cell(row=r, column=2, value=line).font = body_font
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value='Common Mistakes').font = sub_font
    r += 1
    for line in [
        '1. Missing title or price - these are the only required columns.',
        '2. Price in paise instead of rupees - write 249, not 24900.',
        '3. Empty description - Meta requires 10+ chars. Auto-generated if empty.',
        '4. Duplicate retailer_id - leave the "id" column empty, system generates it.',
    ]:
        ws.cell(row=r, column=2, value=line).font = body_font
        r += 1


def build_column_map_sheet(wb):
    ws = wb.create_sheet('Column Map (Dev Reference)')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35

    hdr_font = Font(bold=True, size=10)
    for ci, hdr in enumerate(['Spreadsheet Column', 'Internal DB Field', 'Meta API Field', 'Notes'], 1):
        ws.cell(row=1, column=ci, value=hdr).font = hdr_font
        ws.cell(row=1, column=ci).fill = LIGHT_GRAY

    # META_COLUMN_ALIASES mapping
    mapping = [
        ('id', 'retailer_id', 'retailer_id', 'Auto-generated: {branch}-{item}-{size}'),
        ('title', 'name', 'title', 'REQUIRED'),
        ('description', 'description', 'description', 'Min 10 chars for Meta'),
        ('price', 'price_paise (via parser)', 'price', 'Stored as paise, synced as "N.NN INR"'),
        ('sale_price', 'sale_price_paise', 'sale_price', 'Optional discounted price'),
        ('image_link', 'image_url', 'image_link', 'HTTPS URL, fallback placeholder used'),
        ('availability', 'is_available (boolean)', 'availability', '"in stock" / "out of stock"'),
        ('brand', 'brand', 'brand', 'Falls back to restaurant name'),
        ('link', 'link', 'link', 'Auto-generated product URL'),
        ('item_group_id', 'item_group_id', 'item_group_id', 'Auto: {branch-slug}-{item-slug}'),
        ('size', 'size + variant_value', 'size', 'Triggers variant grouping'),
        ('condition', '(ignored)', 'condition', 'Always "new" for food'),
        ('google_product_category', 'google_product_category', 'google_product_category', 'Default: Food, Beverages & Tobacco > Food Items'),
        ('fb_product_category', 'fb_product_category', 'fb_product_category', 'Default: Food & Beverages > Prepared Food'),
        ('product_tags[0]', 'product_tags[]', 'product_tags[0]', 'Diet tag: Veg/Non-Veg/Egg/Jain/Vegan'),
        ('product_tags[1]', 'product_tags[]', 'product_tags[1]', 'Secondary tag'),
        ('category', 'category_id (looked up)', '(not synced)', 'Auto-creates if new'),
        ('food_type', 'food_type', '(not synced)', 'veg/non_veg/egg/vegan. Default: veg'),
        ('is_bestseller', 'is_bestseller', '(not synced)', 'true/false'),
        ('branch', '(routes to branch_id)', '(not synced)', 'Multi-branch column. Also: outlet, location'),
        ('custom_label_3', '(routes to branch_id)', '(not synced)', 'Alt branch slug column'),
        ('custom_label_2', '(routes to branch_id)', '(not synced)', 'Alt branch area column'),
    ]
    for ri, row_data in enumerate(mapping, 2):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val).font = Font(size=9)


def main():
    wb = Workbook()
    build_menu_sheet(wb)
    build_instructions_sheet(wb)
    build_column_map_sheet(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f'Generated: {OUTPUT_PATH}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Menu items: {len(ITEMS)} rows')
    print(f'Columns: {len(COLUMNS)}')


if __name__ == '__main__':
    main()
